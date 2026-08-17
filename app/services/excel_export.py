"""Excel export — DESIGN.md §4, implementation plans #3 Task 3 and #5 Task 5.

Produces a five-sheet workbook filtered to a date range and to current
(non-superseded), non-deleted versions only:

  * **Sales** — one row per sale item, with the sale's payment split.
  * **Expenses** — description, amount, payment split, "linked to restock"
    badge (resolved via ``find_batch_for_expense`` — never re-derived).
  * **Debts** — customer, note, total, paid, outstanding, status.
  * **Balance summary** — range totals computed via ``domain.balance`` so the
    workbook can never disagree with what the app shows for the same range.
  * **Auditoría** (plan-05 Task 5.1) — every audit event in the range, all
    categories and change types, with the same summary sentences the screen
    shows.

``export_audit_events`` (plan-05 Task 5.2) writes a single-sheet workbook
using the Auditoría screen's **exact current filters**, for handing over
"everything user X touched in period Y".

Money is stored as integer cents in the DB and converted to decimal currency
figures here, at the export boundary. Working directory: no writes to the
ledger; this is a pure read against the passed ``conn``.
"""

from __future__ import annotations

import sqlite3
from datetime import date, datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from app.db.repositories import batches as batches_repo
from app.db.repositories import expenses as expenses_repo
from app.db.repositories import sales as sales_repo
from app.domain import audit as audit_domain
from app.domain.balance import compute_range_totals
from app.ui import strings_es
from app.ui.views import audit_controller

_MONEY_FORMAT = "#,##0.00"
_CENTS_PER_UNIT = 100
_HEADER_FONT = Font(bold=True)

_SALES_HEADERS = [
    strings_es.EXPORT_COL_DATE,
    strings_es.EXPORT_COL_PRODUCT,
    strings_es.EXPORT_COL_QUANTITY,
    strings_es.EXPORT_COL_UNIT_PRICE,
    strings_es.EXPORT_COL_LINE_TOTAL,
    strings_es.EXPORT_COL_CASH,
    strings_es.EXPORT_COL_QR,
    strings_es.EXPORT_COL_CREDIT,
    strings_es.EXPORT_COL_CUSTOMER,
    strings_es.EXPORT_COL_NOTE,
]

_EXPENSES_HEADERS = [
    strings_es.EXPORT_COL_DATE,
    strings_es.EXPORT_COL_DESCRIPTION,
    strings_es.EXPORT_COL_USER,
    strings_es.EXPORT_COL_CASH,
    strings_es.EXPORT_COL_QR,
    strings_es.EXPORT_COL_TOTAL,
    strings_es.EXPORT_COL_LINKED_RESTOCK,
]

_DEBTS_HEADERS = [
    strings_es.EXPORT_COL_CUSTOMER,
    strings_es.EXPORT_COL_NOTE,
    strings_es.EXPORT_COL_TOTAL,
    strings_es.EXPORT_COL_PAID,
    strings_es.EXPORT_COL_OUTSTANDING,
    strings_es.EXPORT_COL_STATUS,
    strings_es.EXPORT_COL_DATE,
]

_AUDIT_HEADERS = [
    strings_es.AUDIT_COL_CATEGORY,
    strings_es.AUDIT_COL_CHANGE_TYPE,
    strings_es.EXPORT_COL_DATE,
    strings_es.EXPORT_COL_USER,
    strings_es.AUDIT_COL_SUMMARY,
]


def _day_bounds(day: date) -> tuple[int, int]:
    """``(start_ts, end_ts)`` covering the whole local day."""
    start = int(datetime(day.year, day.month, day.day).timestamp())
    end = int(datetime(day.year, day.month, day.day, 23, 59, 59).timestamp())
    return start, end


def _cents_to_currency(cents: int) -> float:
    """Convert stored integer cents to a decimal currency figure."""
    return cents / _CENTS_PER_UNIT


def _format_ts(ts: int) -> str:
    return datetime.fromtimestamp(ts).strftime("%d/%m/%Y %H:%M")


def _write_headers(sheet, headers: list[str]) -> None:
    for column, header in enumerate(headers, start=1):
        cell = sheet.cell(row=1, column=column, value=header)
        cell.font = _HEADER_FONT


def _autosize(sheet, widths: list[int]) -> None:
    for index, width in enumerate(widths, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width


def _current_sales_in_range(
    conn: sqlite3.Connection, from_ts: int, to_ts: int
) -> list[sqlite3.Row]:
    """Current, non-deleted sale versions whose version row lies in the range."""
    return conn.execute(
        """
        SELECT s.id, s.logical_id, s.timestamp, s.is_credit, s.customer_name,
               s.customer_note, u.name AS user_name
        FROM sales s
        JOIN (
            SELECT logical_id, MAX(version) AS max_version
            FROM sales
            WHERE timestamp >= ? AND timestamp <= ?
            GROUP BY logical_id
        ) latest
          ON latest.logical_id = s.logical_id
         AND latest.max_version = s.version
        JOIN users u ON u.id = s.current_user
        WHERE s.deleted_at IS NULL
        ORDER BY s.timestamp
        """,
        (from_ts, to_ts),
    ).fetchall()


def _current_expenses_in_range(
    conn: sqlite3.Connection, from_ts: int, to_ts: int
) -> list[sqlite3.Row]:
    """Current, non-deleted expense versions whose version row lies in the range."""
    return conn.execute(
        """
        SELECT e.id, e.logical_id, e.timestamp, e.description, e.amount,
               u.name AS user_name
        FROM expenses e
        JOIN (
            SELECT logical_id, MAX(version) AS max_version
            FROM expenses
            WHERE timestamp >= ? AND timestamp <= ?
            GROUP BY logical_id
        ) latest
          ON latest.logical_id = e.logical_id
         AND latest.max_version = e.version
        JOIN users u ON u.id = e.user_id
        WHERE e.deleted_at IS NULL
        ORDER BY e.timestamp
        """,
        (from_ts, to_ts),
    ).fetchall()


def _current_credit_sales_in_range(
    conn: sqlite3.Connection, from_ts: int, to_ts: int
) -> list[sqlite3.Row]:
    """Current, non-deleted credit sale versions whose version row lies in the range."""
    return conn.execute(
        """
        SELECT s.id, s.logical_id, s.timestamp, s.customer_name, s.customer_note
        FROM sales s
        JOIN (
            SELECT logical_id, MAX(version) AS max_version
            FROM sales
            WHERE timestamp >= ? AND timestamp <= ?
              AND is_credit = 1
            GROUP BY logical_id
        ) latest
          ON latest.logical_id = s.logical_id
         AND latest.max_version = s.version
        WHERE s.deleted_at IS NULL
        ORDER BY s.timestamp
        """,
        (from_ts, to_ts),
    ).fetchall()


def _debt_paid_for_sale(conn: sqlite3.Connection, sale_id: int) -> int:
    row = conn.execute(
        "SELECT COALESCE(SUM(amount), 0) AS total FROM debt_payments"
        " WHERE sale_id = ?",
        (sale_id,),
    ).fetchone()
    return int(row["total"])


def export_range(
    conn: sqlite3.Connection, date_from: date, date_to: date, output_path: str | Path
) -> Path:
    """Write the four-sheet export for ``[date_from, date_to]`` to ``output_path``.

    Raises ``ValueError`` if the range is empty (``date_from > date_to``); the
    authoritative client-side check lives in ``export_controller.validate_range``
    and this stays as a defensive assertion.
    """
    if date_from > date_to:
        raise ValueError("date_from must not be after date_to.")
    from_ts, to_ts = _day_bounds(date_from)[0], _day_bounds(date_to)[1]

    workbook = Workbook()
    _fill_sales_sheet(workbook, conn, from_ts, to_ts)
    _fill_expenses_sheet(workbook, conn, from_ts, to_ts)
    _fill_debts_sheet(workbook, conn, from_ts, to_ts)
    _fill_balance_sheet(workbook, conn, from_ts, to_ts)
    _fill_audit_sheet(workbook, conn, from_ts, to_ts)
    default_sheet = workbook.active  # the empty default sheet, if still present
    if default_sheet is not None:
        workbook.remove(default_sheet)
    workbook.save(str(output_path))
    return Path(output_path)


def _fill_sales_sheet(
    workbook: Workbook, conn: sqlite3.Connection, from_ts: int, to_ts: int
) -> None:
    sheet = workbook.create_sheet(strings_es.EXPORT_SHEET_SALES)
    _write_headers(sheet, _SALES_HEADERS)
    row_index = 2
    for sale in _current_sales_in_range(conn, from_ts, to_ts):
        items = sales_repo.get_sale_items(conn, int(sale["id"]))
        payments = sales_repo.get_sale_payments(conn, int(sale["id"]))
        cash_paid = sum(
            int(p["amount"]) for p in payments if p["method"] == "cash"
        )
        qr_paid = sum(int(p["amount"]) for p in payments if p["method"] == "qr")
        is_credit = bool(sale["is_credit"])
        for item in items:
            total_cents = int(item["quantity"]) * int(item["unit_price_applied"])
            sheet.cell(row=row_index, column=1, value=_format_ts(int(sale["timestamp"])))
            sheet.cell(row=row_index, column=2, value=item["product_name"])
            sheet.cell(row=row_index, column=3, value=int(item["quantity"]))
            sheet.cell(
                row=row_index, column=4, value=_cents_to_currency(int(item["unit_price_applied"]))
            ).number_format = _MONEY_FORMAT
            sheet.cell(row=row_index, column=5, value=_cents_to_currency(total_cents)).number_format = (
                _MONEY_FORMAT
            )
            sheet.cell(row=row_index, column=6, value=_cents_to_currency(cash_paid)).number_format = (
                _MONEY_FORMAT
            )
            sheet.cell(row=row_index, column=7, value=_cents_to_currency(qr_paid)).number_format = (
                _MONEY_FORMAT
            )
            sheet.cell(
                row=row_index,
                column=8,
                value=strings_es.EXPORT_YES if is_credit else strings_es.EXPORT_NO,
            )
            sheet.cell(row=row_index, column=9, value=sale["customer_name"] or "")
            sheet.cell(row=row_index, column=10, value=sale["customer_note"] or "")
            row_index += 1
    _autosize(sheet, [18, 24, 10, 14, 14, 12, 12, 10, 20, 24])


def _fill_expenses_sheet(
    workbook: Workbook, conn: sqlite3.Connection, from_ts: int, to_ts: int
) -> None:
    sheet = workbook.create_sheet(strings_es.EXPORT_SHEET_EXPENSES)
    _write_headers(sheet, _EXPENSES_HEADERS)
    row_index = 2
    for expense in _current_expenses_in_range(conn, from_ts, to_ts):
        payments = expenses_repo.get_expense_payments(conn, int(expense["id"]))
        cash_paid = sum(int(p["amount"]) for p in payments if p["method"] == "cash")
        qr_paid = sum(int(p["amount"]) for p in payments if p["method"] == "qr")
        linked = (
            batches_repo.find_batch_for_expense(conn, int(expense["logical_id"]))
            is not None
        )
        sheet.cell(row=row_index, column=1, value=_format_ts(int(expense["timestamp"])))
        sheet.cell(row=row_index, column=2, value=expense["description"])
        sheet.cell(row=row_index, column=3, value=expense["user_name"])
        sheet.cell(row=row_index, column=4, value=_cents_to_currency(cash_paid)).number_format = (
            _MONEY_FORMAT
        )
        sheet.cell(row=row_index, column=5, value=_cents_to_currency(qr_paid)).number_format = (
            _MONEY_FORMAT
        )
        sheet.cell(row=row_index, column=6, value=_cents_to_currency(int(expense["amount"]))).number_format = (
            _MONEY_FORMAT
        )
        sheet.cell(
            row=row_index,
            column=7,
            value=strings_es.EXPORT_YES if linked else strings_es.EXPORT_NO,
        )
        row_index += 1
    _autosize(sheet, [18, 30, 16, 12, 12, 14, 22])


def _fill_debts_sheet(
    workbook: Workbook, conn: sqlite3.Connection, from_ts: int, to_ts: int
) -> None:
    sheet = workbook.create_sheet(strings_es.EXPORT_SHEET_DEBTS)
    _write_headers(sheet, _DEBTS_HEADERS)
    row_index = 2
    for sale in _current_credit_sales_in_range(conn, from_ts, to_ts):
        items = sales_repo.get_sale_items(conn, int(sale["id"]))
        total_cents = sum(
            int(item["quantity"]) * int(item["unit_price_applied"]) for item in items
        )
        paid_cents = _debt_paid_for_sale(conn, int(sale["id"]))
        outstanding = total_cents - paid_cents
        status = (
            strings_es.EXPORT_STATUS_OPEN
            if outstanding > 0
            else strings_es.EXPORT_STATUS_SETTLED
        )
        sheet.cell(row=row_index, column=1, value=sale["customer_name"] or "")
        sheet.cell(row=row_index, column=2, value=sale["customer_note"] or "")
        sheet.cell(row=row_index, column=3, value=_cents_to_currency(total_cents)).number_format = (
            _MONEY_FORMAT
        )
        sheet.cell(row=row_index, column=4, value=_cents_to_currency(paid_cents)).number_format = (
            _MONEY_FORMAT
        )
        sheet.cell(row=row_index, column=5, value=_cents_to_currency(outstanding)).number_format = (
            _MONEY_FORMAT
        )
        sheet.cell(row=row_index, column=6, value=status)
        sheet.cell(row=row_index, column=7, value=_format_ts(int(sale["timestamp"])))
        row_index += 1
    _autosize(sheet, [22, 26, 14, 14, 14, 14, 18])


def _fill_audit_sheet(
    workbook: Workbook, conn: sqlite3.Connection, from_ts: int, to_ts: int
) -> None:
    """Fifth sheet (plan-05 Task 5.1): every audit event in the range, across
    all categories and change types — a complete record for that period,
    consistent with how the other four sheets are range-only with no extra
    filtering."""
    sheet = workbook.create_sheet(strings_es.EXPORT_SHEET_AUDIT)
    _write_headers(sheet, _AUDIT_HEADERS)
    events = audit_domain.list_audit_events(
        conn, since=from_ts, until=to_ts, limit=None
    )
    _write_audit_rows(sheet, events)
    _autosize(sheet, [18, 16, 18, 16, 70])


def _write_audit_rows(sheet, events) -> None:
    """Write audit events as rows, reusing the screen's summary builders so
    the workbook text matches the on-screen text exactly (plan-05 Task 5)."""
    for index, event in enumerate(events, start=2):
        sheet.cell(
            row=index, column=1, value=audit_controller.category_label(event.category)
        )
        sheet.cell(
            row=index,
            column=2,
            value=audit_controller.change_type_label(event.change_type),
        )
        sheet.cell(row=index, column=3, value=_format_ts(event.timestamp))
        sheet.cell(row=index, column=4, value=event.user_name)
        sheet.cell(row=index, column=5, value=audit_controller.summary_for(event))


def export_audit_events(
    conn: sqlite3.Connection,
    filters: audit_domain.AuditFilters,
    output_path: str | Path,
) -> Path:
    """Dedicated "Exportar esta vista" workbook (plan-05 Task 5.2).

    A single-sheet workbook using the **exact filters currently active** in
    the Auditoría screen (time range, user, categories, change types) — not
    the full unfiltered history — so "show me everything Juan touched in the
    last 7 days" can be handed over as a file. All events matching the
    filters are included (newest first), not just the first page.
    """
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = strings_es.EXPORT_SHEET_AUDIT
    _write_headers(sheet, _AUDIT_HEADERS)
    events = audit_domain.list_audit_events(
        conn,
        since=filters.since,
        until=filters.until,
        user_id=filters.user_id,
        categories=filters.categories,
        change_types=filters.change_types,
        limit=None,
    )
    _write_audit_rows(sheet, events)
    _autosize(sheet, [18, 16, 18, 16, 70])
    workbook.save(str(output_path))
    return Path(output_path)


def _fill_balance_sheet(
    workbook: Workbook, conn: sqlite3.Connection, from_ts: int, to_ts: int
) -> None:
    sheet = workbook.create_sheet(strings_es.EXPORT_SHEET_BALANCE)
    _write_headers(sheet, [strings_es.EXPORT_COL_TOTAL])
    totals = compute_range_totals(conn, from_ts, to_ts)
    labels = [
        strings_es.EXPORT_BALANCE_CASH_SALES,
        strings_es.EXPORT_BALANCE_QR_SALES,
        strings_es.EXPORT_BALANCE_DEBTS_COLLECTED,
        strings_es.EXPORT_BALANCE_CASH_EXPENSES,
        strings_es.EXPORT_BALANCE_QR_EXPENSES,
        strings_es.EXPORT_BALANCE_NET,
    ]
    values = [
        totals["cash_sales"],
        totals["qr_sales"],
        totals["debts_collected"],
        totals["cash_expenses"],
        totals["qr_expenses"],
        totals["net"],
    ]
    for row_index, (label, value) in enumerate(zip(labels, values, strict=True), start=2):
        sheet.cell(row=row_index, column=1, value=label)
        cell = sheet.cell(row=row_index, column=2, value=_cents_to_currency(value))
        cell.number_format = _MONEY_FORMAT
        if label == strings_es.EXPORT_BALANCE_NET:
            cell.font = _HEADER_FONT
    _autosize(sheet, [30, 16])
