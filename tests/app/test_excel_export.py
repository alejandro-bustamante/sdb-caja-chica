"""Unit tests for the Excel export service and its controller (plan-03 Task 7).

The build-smoke tests already cover that the export view constructs; these
tests cover the actual workbook contents and the date-range validation that
the view delegates to ``export_controller``.
"""

from __future__ import annotations

from datetime import date

import pytest
from openpyxl import load_workbook

from app.db.repositories import batches as batches_repo
from app.db.repositories import debts as debts_repo
from app.db.repositories import expenses as expenses_repo
from app.db.repositories import products as products_repo
from app.db.repositories import sales as sales_repo
from app.domain import audit as audit_domain
from app.domain.types import ExpensePaymentInput, SaleItemInput, SalePaymentInput
from app.services import excel_export
from app.ui import strings_es
from app.ui.views import audit_controller, export_controller


def _seed(conn, user_id):
    pid = products_repo.create_product(conn, "Azúcar", 1000, user_id)
    batches_repo.create_batch(
        conn, [(pid, 10)], [ExpensePaymentInput("cash", 5000)], "flete", user_id
    )
    sales_repo.create_sale(
        conn,
        [SaleItemInput(product_id=pid, quantity=2, unit_price_applied=1000)],
        [SalePaymentInput("cash", 2000)],
        False,
        None,
        None,
        user_id,
    )
    sales_repo.create_sale(
        conn,
        [SaleItemInput(product_id=pid, quantity=1, unit_price_applied=1000)],
        [],
        True,
        "Pepe",
        "a cuaderno",
        user_id,
    )
    debts_repo.record_partial_payment(conn, 2, 400, user_id)
    expenses_repo.create_expense(
        conn, "Luz", [ExpensePaymentInput("cash", 3000)], user_id
    )
    return pid


def _load(path):
    return load_workbook(str(path))


def _cents(value: object) -> int:
    """A spreadsheet cell value (int/float/str...) to integer cents."""
    return int(float(value) * 100)  # type: ignore[arg-type]


def test_export_writes_five_sheets(conn, user_id, tmp_path):
    _seed(conn, user_id)
    today = date.today()
    out = excel_export.export_range(conn, today, today, tmp_path / "export.xlsx")

    assert out.exists()
    workbook = _load(out)
    assert set(workbook.sheetnames) == {
        strings_es.EXPORT_SHEET_SALES,
        strings_es.EXPORT_SHEET_EXPENSES,
        strings_es.EXPORT_SHEET_DEBTS,
        strings_es.EXPORT_SHEET_BALANCE,
        strings_es.EXPORT_SHEET_AUDIT,
    }


def test_export_sales_sheet_rows(conn, user_id, tmp_path):
    _seed(conn, user_id)
    out = excel_export.export_range(conn, date.today(), date.today(), tmp_path / "out.xlsx")
    sheet = _load(out)[strings_es.EXPORT_SHEET_SALES]

    assert sheet.cell(row=1, column=2).value == strings_es.EXPORT_COL_PRODUCT
    rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if r[1]]
    assert len(rows) == 2  # one row per sale item
    assert sum(_cents(row[4]) for row in rows) == 3000  # line totals
    assert sum(_cents(row[5]) for row in rows) == 2000  # cash paid


def test_export_debts_sheet_status(conn, user_id, tmp_path):
    _seed(conn, user_id)
    out = excel_export.export_range(conn, date.today(), date.today(), tmp_path / "out.xlsx")
    sheet = _load(out)[strings_es.EXPORT_SHEET_DEBTS]

    rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if r[0]]
    assert rows[0][0] == "Pepe"
    assert _cents(rows[0][2]) == 1000
    assert _cents(rows[0][3]) == 400
    assert _cents(rows[0][4]) == 600
    assert rows[0][5] == strings_es.EXPORT_STATUS_OPEN


def test_export_balance_sheet_net(conn, user_id, tmp_path):
    _seed(conn, user_id)
    out = excel_export.export_range(conn, date.today(), date.today(), tmp_path / "out.xlsx")
    sheet = _load(out)[strings_es.EXPORT_SHEET_BALANCE]

    values = {row[0]: row[1] for row in sheet.iter_rows(min_row=2, values_only=True) if row[0]}
    assert _cents(values[strings_es.EXPORT_BALANCE_CASH_SALES]) == 2000
    assert _cents(values[strings_es.EXPORT_BALANCE_QR_SALES]) == 0
    assert _cents(values[strings_es.EXPORT_BALANCE_DEBTS_COLLECTED]) == 400
    assert _cents(values[strings_es.EXPORT_BALANCE_CASH_EXPENSES]) == 8000
    assert _cents(values[strings_es.EXPORT_BALANCE_NET]) == -5600


def _audit_sheet_rows(out) -> list[tuple]:
    sheet = _load(out)[strings_es.EXPORT_SHEET_AUDIT]
    return [
        r for r in sheet.iter_rows(min_row=2, values_only=True) if any(r)
    ]


def test_export_range_audit_sheet_covers_all_events(conn, user_id, tmp_path):
    """The general export's Auditoría sheet matches what the screen would show
    with "Todo/Todos/Todas" filters for the same date range (plan-05 Task
    5.1 / 5 exit criteria)."""
    _seed(conn, user_id)
    out = excel_export.export_range(conn, date.today(), date.today(), tmp_path / "out.xlsx")

    rows = _audit_sheet_rows(out)
    expected = audit_domain.count_audit_events(conn)
    assert len(rows) == expected == 7
    # Headers are the audit-specific ones.
    sheet = _load(out)[strings_es.EXPORT_SHEET_AUDIT]
    assert sheet.cell(row=1, column=1).value == strings_es.AUDIT_COL_CATEGORY
    assert sheet.cell(row=1, column=2).value == strings_es.AUDIT_COL_CHANGE_TYPE
    assert sheet.cell(row=1, column=5).value == strings_es.AUDIT_COL_SUMMARY
    # Summaries match the screen's controller output for the same events.
    events = audit_domain.list_audit_events(conn, limit=None)
    assert len(rows) == len(events)
    summaries = [audit_controller.summary_for(e) for e in events]
    assert {r[4] for r in rows} == set(summaries)


def test_export_audit_events_filtered_matches_query(conn, user_id, other_user_id, tmp_path):
    """The dedicated "Exportar esta vista" workbook uses the exact active
    filters, including a single user/category/type (plan-05 Task 5.2)."""
    _seed(conn, user_id)
    from app.domain.types import SaleItemInput, SalePaymentInput

    pid = products_repo.create_product(conn, "Pan", 500, other_user_id)
    sales_repo.create_sale(
        conn,
        [SaleItemInput(product_id=pid, quantity=1, unit_price_applied=500)],
        [SalePaymentInput("cash", 500)],
        False,
        None,
        None,
        other_user_id,
    )

    filters = audit_domain.AuditFilters(
        user_id=other_user_id,
        categories=(audit_domain.CATEGORY_SALES, audit_domain.CATEGORY_CATALOG),
        change_types=(audit_domain.CHANGE_REGISTRO,),
    )
    out = excel_export.export_audit_events(conn, filters, tmp_path / "filtered.xlsx")
    assert out.exists()

    sheet = _load(out)[strings_es.EXPORT_SHEET_AUDIT]
    rows = [r for r in sheet.iter_rows(min_row=2, values_only=True) if any(r)]
    expected = audit_domain.list_audit_events(
        conn,
        user_id=other_user_id,
        categories=filters.categories,
        change_types=filters.change_types,
        limit=None,
    )
    assert len(rows) == len(expected) == 2  # product creation + sale
    for row, event in zip(rows, expected, strict=True):
        assert row[0] == audit_controller.category_label(event.category)
        assert row[1] == audit_controller.change_type_label(event.change_type)
        assert row[4] == audit_controller.summary_for(event)


def test_export_rejects_reversed_range(conn, user_id, tmp_path):
    with pytest.raises(ValueError):
        excel_export.export_range(conn, date(2024, 1, 2), date(2024, 1, 1), tmp_path / "x.xlsx")


def test_controller_validate_range():
    assert export_controller.validate_range("01/02/2024", "10/02/2024") == (
        date(2024, 2, 1),
        date(2024, 2, 10),
        None,
    )
    date_from, date_to, error = export_controller.validate_range("11", "10/02/2024")
    assert error == strings_es.EXPORT_INVALID_DATE
    date_from, date_to, error = export_controller.validate_range("11/02/2024", "10/02/2024")
    assert error == strings_es.EXPORT_RANGE_ERROR
    from_, _, error = export_controller.validate_range("", "")
    assert error == strings_es.EXPORT_INVALID_DATE


@pytest.mark.parametrize(
    "typed, expected",
    [
        ("10/09/2026", date(2026, 9, 10)),
        ("10/09/26", date(2026, 9, 10)),
        ("10-09-2026", date(2026, 9, 10)),
        ("10.09.26", date(2026, 9, 10)),
        ("10 09 2026", date(2026, 9, 10)),
        ("1/2/14", date(2014, 2, 1)),
        (" 10 / 09 / 2026 ", date(2026, 9, 10)),
        ("10/09/26/", date(2026, 9, 10)),
    ],
)
def test_controller_parse_date_tolerant(typed, expected):
    assert export_controller.parse_date_text(typed) == expected


@pytest.mark.parametrize(
    "typed",
    ["", "   ", "10/09", "10/09/026", "31/02/2026", "foo", "10/a/b"],
)
def test_controller_parse_date_rejects_garbage(typed):
    assert export_controller.parse_date_text(typed) is None


def test_controller_default_file_name():
    name = export_controller.default_file_name(date(2024, 2, 1), date(2024, 2, 10))
    assert "01-02-2024" in name
    assert "10-02-2024" in name
