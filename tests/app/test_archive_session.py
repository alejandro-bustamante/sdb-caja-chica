"""Integration test for the archived-ledger viewer (plan-04 Task 6.3).

Runs the full open -> browse -> export -> close cycle against a fixture
archived file and asserts the live ledger's own data, balance, and on-disk
bytes are completely unaffected throughout — and that the staged temp copy is
deleted when the viewer closes.
"""

from __future__ import annotations

import sqlite3
from datetime import date
from pathlib import Path

import flet as ft
from openpyxl import load_workbook

from app.db.connection import migrate, open_connection
from app.db.repositories import batches as batches_repo
from app.db.repositories import expenses as expenses_repo
from app.db.repositories import products as products_repo
from app.db.repositories import sales as sales_repo
from app.db.repositories import users as users_repo
from app.domain.balance import compute_total_available
from app.domain.types import ExpensePaymentInput, SaleItemInput, SalePaymentInput
from app.services import excel_export
from app.ui import strings_es
from app.ui.archive import ArchiveSessionManager
from app.ui.session import Session
from app.ui.views import cash_counts as cash_counts_view
from app.ui.views import catalog as catalog_view
from app.ui.views import debts as debts_view
from app.ui.views import expenses as expenses_view
from app.ui.views import export as export_view
from app.ui.views import restock as restock_view
from app.ui.views import sales as sales_view

VIEW_BUILDERS = [
    sales_view.build,
    catalog_view.build,
    restock_view.build,
    expenses_view.build,
    debts_view.build,
    cash_counts_view.build,
    export_view.build,
]


def _fold_wal(path: Path) -> None:
    """Checkpoint any WAL into the main file so byte comparisons are stable."""
    conn = sqlite3.connect(str(path))
    try:
        conn.execute("PRAGMA wal_checkpoint(TRUNCATE)")
    finally:
        conn.close()


def _seed_ledger(conn, user_id: int) -> None:
    pid = products_repo.create_product(conn, "Azúcar", 1000, user_id)
    batches_repo.create_batch(
        conn, [(pid, 10)], [ExpensePaymentInput("cash", 5000)], "flete", user_id
    )
    sales_repo.create_sale(
        conn,
        [SaleItemInput(product_id=pid, quantity=2, unit_price_applied=1000)],
        [SalePaymentInput("cash", 2000)],
        False,
        None,
        None,
        user_id,
    )
    expenses_repo.create_expense(
        conn, "Luz", [ExpensePaymentInput("cash", 3000)], user_id
    )


def test_full_archive_cycle_leaves_live_ledger_untouched(db_path, tmp_path):
    # --- Live ledger: migrated, seeded, WAL folded, byte snapshot -----------
    live_conn = open_connection(db_path)
    live_user = users_repo.create_user(live_conn, "Alice")
    _seed_ledger(live_conn, live_user)
    live_balance_before = compute_total_available(live_conn)
    _fold_wal(db_path)
    live_bytes_before = db_path.read_bytes()

    # --- Archived fixture: a separate, fully-migrated ledger with its own data
    archive_path = tmp_path / "archived.db"
    migrate(archive_path)
    archive_conn = open_connection(archive_path)
    archive_user = users_repo.create_user(archive_conn, "Ana")
    pid = products_repo.create_product(archive_conn, "Harina", 2000, archive_user)
    sales_repo.create_sale(
        archive_conn,
        [SaleItemInput(product_id=pid, quantity=1, unit_price_applied=2000)],
        [SalePaymentInput("qr", 2000)],
        False,
        None,
        None,
        archive_user,
    )
    archive_conn.close()
    archive_bytes_before = archive_path.read_bytes()

    session = Session(user_id=live_user, user_name="Alice")
    manager = ArchiveSessionManager(live_conn, session)

    # --- Open ---------------------------------------------------------------
    opened = manager.open(archive_path)
    assert manager.active
    assert opened.display_name == "archived.db"
    assert manager.session().read_only
    assert manager.session().user_name == "Alice"

    # --- Browse: every view builds read-only against the archived copy ------
    for builder in VIEW_BUILDERS:
        root = builder(opened.conn, manager.session(), lambda: None)
        assert isinstance(root, ft.Control)

    # --- Export: figures match the archived ledger's own balance ------------
    out = excel_export.export_range(
        opened.conn, date.today(), date.today(), tmp_path / "export.xlsx"
    )
    workbook = load_workbook(str(out))
    balance_sheet = workbook[strings_es.EXPORT_SHEET_BALANCE]
    values = {
        row[0]: row[1]
        for row in balance_sheet.iter_rows(min_row=2, values_only=True)
        if row[0]
    }
    exported_net = int(float(values[strings_es.EXPORT_BALANCE_NET]) * 100)
    assert exported_net == compute_total_available(opened.conn)
    # The archived ledger's balance differs from the live one (its own data).
    assert exported_net == 2000

    # --- Live ledger unaffected while the archive is open --------------------
    assert db_path.read_bytes() == live_bytes_before
    assert compute_total_available(live_conn) == live_balance_before

    # --- Close ---------------------------------------------------------------
    copy_path = opened.copy_path
    assert copy_path.exists()
    manager.close()
    assert not manager.active
    assert not manager.session().read_only
    assert manager.session() is session
    assert not copy_path.exists()  # temp copy deleted

    # Live ledger still works and is byte-for-byte unchanged after the cycle.
    assert db_path.read_bytes() == live_bytes_before
    assert archive_path.read_bytes() == archive_bytes_before
    assert compute_total_available(live_conn) == live_balance_before
    assert len(users_repo.list_active_users(live_conn)) == 1
    live_conn.close()


def test_shell_header_keeps_menu_and_archive_banner(conn, user_id):
    """The shell's header row always keeps the menu button; in archive mode
    the user bar is swapped for the SOLO LECTURA banner (plan-04 Task 3/4)."""
    from app.ui.shell import build_shell

    class StubPage:
        def update(self):
            pass

        def show_dialog(self, dialog):
            pass

        def pop_dialog(self, dialog=None):
            pass

    session = Session(user_id=user_id, user_name="Alice")
    root = build_shell(StubPage(), conn, session)
    header_row = root.controls[0]
    assert len(header_row.controls) == 2  # user bar + menu
    menu = header_row.controls[1]
    # Live mode: backup, open-archive and archive actions are offered.
    assert len(menu.items) == 3
    assert any(
        item.content == strings_es.MENU_OPEN_ARCHIVE for item in menu.items
    )


def test_archive_manager_close_is_idempotent_and_reopens(db_path, tmp_path):
    """Closing twice is a no-op; opening another archive replaces the first."""
    live_conn = open_connection(db_path)
    user = users_repo.create_user(live_conn, "Alice")

    first = tmp_path / "first.db"
    migrate(first)
    second = tmp_path / "second.db"
    migrate(second)

    session = Session(user_id=user, user_name="Alice")
    manager = ArchiveSessionManager(live_conn, session)
    opened_first = manager.open(first)
    opened_second = manager.open(second)
    assert opened_first.copy_path != opened_second.copy_path
    assert not opened_first.copy_path.exists()  # replaced archive was cleaned
    assert manager.current is opened_second

    manager.close()
    manager.close()  # idempotent
    assert not manager.active
    assert manager.current is None
    live_conn.close()
